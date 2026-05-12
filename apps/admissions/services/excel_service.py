from openpyxl import Workbook
from openpyxl.styles import Font

from django.http import HttpResponse

from apps.admissions.models import (
    Application
)


def export_enrollments_excel():

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = 'Enrollments'

    headers = [
        'ID',
        'ФИО',
        'Специальность',
        'GPA',
        'Exam Score',
        'Total Score',
        'Бюджет',
        'Статус',
    ]

    for col_num, header in enumerate(
        headers,
        1
    ):

        cell = sheet.cell(
            row=1,
            column=col_num
        )

        cell.value = header

        cell.font = Font(
            bold=True
        )

    applications = (
        Application.objects.select_related(
            'applicant',
            'speciality'
        ).filter(
            is_recommended=True
        )
    )

    row_num = 2

    for app in applications:

        sheet.cell(
            row=row_num,
            column=1
        ).value = app.id

        sheet.cell(
            row=row_num,
            column=2
        ).value = (
            f'{app.applicant.last_name} '
            f'{app.applicant.first_name}'
        )

        sheet.cell(
            row=row_num,
            column=3
        ).value = app.speciality.title

        sheet.cell(
            row=row_num,
            column=4
        ).value = float(
            app.applicant.gpa
        )

        sheet.cell(
            row=row_num,
            column=5
        ).value = app.exam_score

        sheet.cell(
            row=row_num,
            column=6
        ).value = app.total_score

        sheet.cell(
            row=row_num,
            column=7
        ).value = (
            'БЮДЖЕТ'
            if app.is_budget
            else 'ПЛАТНОЕ'
        )

        sheet.cell(
            row=row_num,
            column=8
        ).value = app.status

        row_num += 1

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-'
            'officedocument.spreadsheetml.sheet'
        )
    )

    response[
        'Content-Disposition'
    ] = (
        'attachment; '
        'filename=enrollments.xlsx'
    )

    workbook.save(response)

    return response